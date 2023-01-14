from __future__ import annotations
from typing import Any
from abc import ABC, abstractmethod
from copy import copy

import openpyxl

ALPH = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
COLS = list(ALPH) + [
    a+b
    for a in ALPH
    for b in ALPH
]

class TableConfig:
    MAJOR_DISPLAY_ROW = 1
    MINOR_DISPLAY_ROW = 2
    COLUMN_NAME_ROW = 3
    VALUES_BEGIN_ROW = 4

class Workbook(ABC):

    WORKBOOK_DRIVE = "drive"
    WORKBOOK_EXCEL = "excel"

    def __init__(self, book_options) -> None:
        self.book_options = book_options

    @classmethod
    def from_drive(self, drive_link, options={}):
        options.update({"link": drive_link})
        return DriveBook(options)

    @classmethod
    def from_xlsx(self, filepath, options={}):
        options.update({"path": filepath})
        return ExcelBook(options)

    @classmethod
    def from_options(self, options, link_name="link", path_name="path", sheet_name="sheet"):
        if link_name in options:
            wb = Workbook.from_drive(options[link_name], copy(options))
        elif path_name in options:
            wb =  Workbook.from_xlsx(options[path_name], copy(options))
        else:
            raise ValueError(f"Excepted either {link_name} or {path_name} in object {options}.")
        wb.set_worksheet(options[sheet_name])
        return wb

    @abstractmethod
    def create_worksheet(self, worksheet_name):
        pass

    @abstractmethod
    def set_worksheet(self, worksheet_name):
        pass

    @abstractmethod
    def row_values(self, row_idx: int):
        pass

    @abstractmethod
    def col_values(self, col_idx: int):
        pass

    @abstractmethod
    def update_values(self, row_start_idx: int, row_end_idx: int, col_start_idx: int, col_end_idx: int, values: list[list[Any]]):
        pass

    @abstractmethod
    def get_values(self, row_start_idx: int, row_end_idx: int, col_start_idx: int, col_end_idx: int) -> list[list[Any]]:
        pass

    @abstractmethod
    def close(self):
        pass

    @abstractmethod
    def clear(self):
        pass

class DriveBook(Workbook):

    def __init__(self, book_options) -> None:
        super().__init__(book_options)
        import gspread
        from oauth2client.service_account import ServiceAccountCredentials

        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive',
        ]

        credentials = ServiceAccountCredentials.from_json_keyfile_name('secrets/credentials.json', scope)
        client = gspread.authorize(credentials)
        self.workbook = client.open_by_url(book_options["link"])
        self.worksheet = None

    def create_worksheet(self, worksheet_name):
        self.worksheet = self.workbook.add_worksheet(worksheet_name, 1000, 100)

    def set_worksheet(self, worksheet_name_or_id):
        import gspread
        try:
            if isinstance(worksheet_name_or_id, str):
                self.worksheet = self.workbook.worksheet(worksheet_name_or_id)
            else:
                self.worksheet = self.workbook.get_worksheet(worksheet_name_or_id)
        except gspread.WorksheetNotFound:
            self.create_worksheet(worksheet_name_or_id)

    def row_values(self, row_idx: int):
        return self.worksheet.row_values(row_idx)

    def col_values(self, col_idx: int):
        return self.worksheet.col_values(col_idx)

    def update_values(self, row_start_idx: int, row_end_idx: int, col_start_idx: int, col_end_idx: int, values: list[list[Any]]):
        range_start = COLS[col_start_idx] + str(row_start_idx + 1)
        range_end = COLS[col_end_idx] + str(row_end_idx + 1)
        return self.worksheet.update(f"{range_start}:{range_end}", values)

    def get_values(self, row_start_idx: int, row_end_idx: int, col_start_idx: int, col_end_idx: int) -> list[list[Any]]:
        range_start = COLS[col_start_idx] + str(row_start_idx + 1)
        range_end = COLS[col_end_idx] + str(row_end_idx + 1)
        return self.worksheet.get(f"{range_start}:{range_end}")

    def close(self):
        return

    def clear(self):
        return self.worksheet.clear()

class ExcelBook(Workbook):

    def __init__(self, book_options) -> None:
        super().__init__(book_options)
        self.workbook = openpyxl.load_workbook(book_options["path"], book_options.get("read_only", False), keep_vba=book_options.get("vba_only", False), data_only=book_options.get("data_only", True), keep_links=book_options.get("keep_links", False))
        self.worksheet = None

    def create_worksheet(self, worksheet_name):
        self.worksheet = self.workbook.create_sheet(worksheet_name)

    def set_worksheet(self, worksheet_name_or_id):
        try:
            if isinstance(worksheet_name_or_id, str):
                self.worksheet = self.workbook[worksheet_name_or_id]
            else:
                self.worksheet = self.workbook.worksheets[worksheet_name_or_id]
        except KeyError:
            self.create_worksheet(worksheet_name_or_id)

    def row_values(self, row_idx: int):
        vals = [
            cell.value
            for cell in self.worksheet[str(row_idx)]
        ]
        while vals and vals[-1] is None:
            vals.pop()
        return vals

    def col_values(self, col_idx: int):
        vals = [
            cell.value
            for cell in self.worksheet[COLS[col_idx-1]]
        ]
        while vals and vals[-1] is None:
            vals.pop()
        return vals

    def update_values(self, row_start_idx: int, row_end_idx: int, col_start_idx: int, col_end_idx: int, values: list[list[Any]]):
        for i, row in enumerate(range(row_start_idx+1, row_end_idx+2)):
            for j, col in enumerate(range(col_start_idx+1, col_end_idx+2)):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = values[i][j]

    def get_values(self, row_start_idx: int, row_end_idx: int, col_start_idx: int, col_end_idx: int) -> list[list[Any]]:
        data = [
            [
                None
                for _ in range(col_start_idx+1, col_end_idx+2)
            ]
            for _ in range(row_start_idx+1, row_end_idx+2)
        ]
        for i, row in enumerate(range(row_start_idx+1, row_end_idx+2)):
            for j, col in enumerate(range(col_start_idx+1, col_end_idx+2)):
                cell = self.worksheet.cell(row=row, column=col)
                data[i][j] = cell.value
        return data

    def close(self):
        self.workbook.save(self.book_options["path"])

    def clear(self):
        name = self.worksheet.title
        self.workbook.remove(self.worksheet)
        self.create_worksheet(name)
